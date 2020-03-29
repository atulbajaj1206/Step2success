from pywinauto.application import Application
import time
import pandas as pd
import time
import openpyxl
import sys

jumpserver=''
login=''
password=''
gsk_login=''
gsk_password=''


############################################################

app = Application ().Start (cmd_line=u'putty -ssh {}@{}'.format(login,jumpserver))
putty = app.PuTTY
putty.wait ('ready')
time.sleep (8)
putty.type_keys (password)
putty.type_keys ("{ENTER}")
time.sleep (10)
print("Enterning NMSTEL password")
putty.type_keys (password)
putty.type_keys("{ENTER}")
time.sleep (3)

def UpdateRoutersResultsToSheet_fping(routers,fping,counter):
	
    workbook_obj = openpyxl.load_workbook('input.xlsx')
    print('saving excel')
    sheet_obj = workbook_obj['Sheet1']
    i=counter+2
    name = routers
    sheet_obj.cell(row=i, column=2).value = fping
    
    workbook_obj.save('input.xlsx')

    '''for i in range(2, sheet_obj.max_row, 1):
      print(sheet_obj.cell(row=i, column=1).value)
           
      if routers == (sheet_obj.cell(row=i, column=1).value):
       
        sheet_obj.cell(row=i, column=3).value = input1
        sheet_obj.cell(row=i, column=4).value = output
        workbook_obj.save('input.xlsx')
        break'''
        
    print('Excel save')
def UpdateRoutersResultsToSheet(routers,input1,output,counter):
	
    workbook_obj = openpyxl.load_workbook('input.xlsx')
    print('saving excel')
    sheet_obj = workbook_obj['Sheet1']
    i=counter+2
    name = routers
    sheet_obj.cell(row=i, column=3).value = input1
    sheet_obj.cell(row=i, column=4).value = output
    workbook_obj.save('input.xlsx')

    '''for i in range(2, sheet_obj.max_row, 1):
      print(sheet_obj.cell(row=i, column=1).value)
           
      if routers == (sheet_obj.cell(row=i, column=1).value):
       
        sheet_obj.cell(row=i, column=3).value = input1
        sheet_obj.cell(row=i, column=4).value = output
        workbook_obj.save('input.xlsx')
        break'''
        
    print('Excel save')

def read():
	f=open('log/putty.log','r')
	lines=f.read()
	#lines=lines.split('{}'.format(router))[-1]
	return(lines)

def refresh():
	f1=open('log/putty.log','w')


def connect(command,router,counter):
	refresh()
	putty.type_keys ("fping {}".format(router), with_spaces=True)
	putty.type_keys("{ENTER}")
	time.sleep (1)
	a=read()
	
	
	if  'alive' in a:
		print('alive')
		UpdateRoutersResultsToSheet_fping(router,'reachable',counter)
		print('if alive')
		
	else:
		time.sleep(2)
		a=read()
		if  'alive' in a:
			
			#print('alive')
			UpdateRoutersResultsToSheet_fping(router,'reachable',counter)
			print('else alive')
		else:
			UpdateRoutersResultsToSheet_fping(router,'Unreachable',counter)
			return()


	putty.type_keys ("l {}".format(router), with_spaces=True)
	putty.type_keys("{ENTER}")
	time.sleep (4)
	a=read()
	c=0
	while not 'Device connected' in a:
		c+=1
		a=read()
		time.sleep (0.5)
		print('.')
		if 'Please enter login credentials' in a:
			refresh()
			putty.type_keys ("{}".format(gsk_login), with_spaces=True)
			print("Entering Gsk logins")
			putty.type_keys("{ENTER}")
			time.sleep (3)
			putty.type_keys ("{}".format(gsk_password), with_spaces=True)
			
			putty.type_keys("{ENTER}")
			time.sleep (3)
			c=0
			
			#time.sleep(5)
			a=read()

		if c>100:
			#putty.type_keys("{CTRL}{Z}")
			#putty.type_keys ("{ENTER}")
			break

	if 'Device connected' in a:
		time.sleep(0.5)


		putty.type_keys ("{}".format(command), with_spaces=True)
		print("Devices connected firing command")
		putty.type_keys("{ENTER}")
		time.sleep (1)
		result=read()
		c2=0
		while 'input rate' not in result:
			print('..')
			time.sleep(0.2)
			result=read()
			c2+=1
			if 'Invalid input detected' in result:
				putty.type_keys ("{}".format("exit"), with_spaces=True)
				putty.type_keys("{ENTER}")
				break

		result=str(result).strip()
		input1=result.split('input rate')[-1]
		input1=input1.split(",")[0]
		output=result.split('output rate')[-1]
		output=output.split(",")[0]
		input1=input1.replace('bits/sec','')
		output=output.replace('bits/sec','')
		print(input1,output)
		UpdateRoutersResultsToSheet(router,input1,output,counter)
		refresh()
		
		putty.type_keys ("exit", with_spaces=True)
		putty.type_keys("{ENTER}")

		a=read()
		while not "close" in a:
			time.sleep(0.5)
			print("__")
			a=read()
		time.sleep(0.5)
		

		
	














df = pd.read_excel('input.xlsx')
devices=df['switch']
command=df['command']
input=df['input']

counter=-1








for i in devices:
  counter+=1
  i=i.strip()
  print("===================")
  print('Router',i,counter)
  if not 'nan'in str(input[counter]):
  	continue
  
  try:
    connect(command[counter ],i,counter)
    
  except Exception as e:
    print("Error",e)