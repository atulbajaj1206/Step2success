import pandas as pd
import time
import os
import openpyxl

#another library for excel which work on cell basis not like pandas whole sheet basis
file_name='input.xlsx'


def excel_write(col_a,col_b,line_no):
	
    workbook_obj = openpyxl.load_workbook(file_name)
   
    sheet_obj = workbook_obj['Sheet1']
    line_no=line_no+2
    #beacuse our counter is 0 and excel starts from 2,1 is header

    
    sheet_obj.cell(row=line_no, column=1).value = 'ankit'

    read_data=(sheet_obj.cell(row=line_no, column=1).value)
    



  
    #pointing row no nad col no
   
    sheet_obj.cell(row=line_no, column=2).value = col_b
    
    workbook_obj.save(file_name)





for j in range(0,10):
	excel_write('step2success',j*j,j)

#opening the file at end of the program
os.system(file_name)

