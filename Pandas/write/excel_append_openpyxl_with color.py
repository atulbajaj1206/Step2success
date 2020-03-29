import pandas as pd
import time
import openpyxl
from openpyxl.styles import PatternFill
import os

file_name='input.xlsx'
def excel_write(col_a,col_b,line_no):

	#loading excel
    workbook_obj = openpyxl.load_workbook(file_name)
   
    sheet_obj = workbook_obj['Sheet1']
    line_no=line_no+2#excel column starts from 2,1 is header
    
    sheet_obj.cell(row=line_no, column=1).value = col_a
    #example if inout no is even then color the cell given code in hex format pls use web for hex color code
    if (line_no%2==0):
    	sheet_obj.cell(row=line_no, column=2).fill = PatternFill(start_color='DCFFDC', end_color='DCFFDC', fill_type = "solid")
   
    sheet_obj.cell(row=line_no, column=2).value = col_b
    
    workbook_obj.save(file_name)
    #save and close file

for j in range(0,50):
	excel_write('step2success',j,j)
	print('writting',j)

#opeing file at the end
os.system(file_name)
