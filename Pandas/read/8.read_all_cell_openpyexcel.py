import pandas as pd
import time
import os
import openpyxl

#another library for excel which work on cell basis not like pandas whole sheet basis
file_name='input.xlsx'



	
workbook_obj = openpyxl.load_workbook(file_name)
   
sheet_obj = workbook_obj['Sheet1']
    
#max_row - count  
for i in range(2, sheet_obj.max_row, 1):
	for j in range(2, sheet_obj.max_column, 1):
		read_data=(sheet_obj.cell(row=i, column=j).value)
		print(read_data)
		print('row:',i,'col:',j)
	print('------------------')

